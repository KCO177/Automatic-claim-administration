import psycopg2
import psycopg2 as pg
import pandas as pd
from qualita_db_dictionary import MyDict as Dic
from datetime import date, timedelta
from openpyxl import load_workbook
#from openpyxl.drawing.image import Image
import datetime
import os
#from mail_attachment import Picture_resizing_for_qa

class Time:

    def current_cw():
        #import datetime
        today = datetime.date.today()
        calendar_week = today.isocalendar()[1]
        # print('today',today, 'CW', calendar_week)
        return calendar_week

    def date_format(date):
        formated_date = date.strftime('%Y - %m - %d')
        return formated_date
        #datetime.date(2023, 3, 22)
        #to format 2023 - 03 - 09
class SetGetData:
    '''Class for setting connection, inserting new claims, simple visualisation and export of all table'''

    def connect_claim():
        # connecting to the db Qualita_test claim table

        try:

            conn = pg.connect(
                database = 'Qualita_test',
                user = 'postgres',
                password = 'kco177',
                host = 'localhost',
                port = '5432')

            # create cursor object
            cur = conn.cursor()

        except (Exception, psycopg2.DatabaseError) as error:

            print ('Error while creating PostgreSQL table', error)

        return conn, cur

    def fetch_data():
        #fetch all data from claim table
        conn, cur = SetGetData.connect_claim()

        try:
            cur.execute('SELECT * FROM claims')

        except:
            print('error !')

        #store the results in data
        data = cur.fetchall()
        return data

    def insert_claim(claim_date, project, part_name, part_id, customer, customer_claim_id, type_of_claim, claimed_failure, amount_of_claimed, stock):
        conn, cur = SetGetData.connect_claim()
        id = (SetGetData.get_last_id())+1
        try:
            cur.execute('INSERT INTO claims VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                        (id, claim_date, project, part_name, part_id, customer, customer_claim_id, type_of_claim, claimed_failure, amount_of_claimed, stock))

        except Exception as e:
            print('error', e)

        # comiting the transaction
        conn.commit()

        # >>>

    def insert_decoded(data):
        conn, cur = SetGetData.connect_claim()
        #id = (SetGetData.get_last_id())+1
        #[3, 'P8', 'SMRC', 'D0323116', 'CH LHD', '04-24-2023', '07:24:37']
        claim_id = data[0]
        project = data[1]
        customer = data[2]
        part_name = data[4]
        part_id = data[3]
        prod_time = data[6]
        prod_day = data[5]
        image = 'image'
        line = 0
        try:
           cur.execute('INSERT INTO decoded VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)',(claim_id, image, project, customer, part_name, part_id, prod_time, prod_day, line))
           print('data in db_decoded inserted')
        except Exception as e:
            print('error', e)
        # comiting the transaction
        conn.commit()

    def get_last_id():
        # return last id vlaue from claim table
        conn, cur = SetGetData.connect_claim()

        try:
            cur.execute('SELECT max(id) FROM claims')

        except:
            print('error in id!')

        #store the results in data
        data = cur.fetchall()
        value = data[0][0]
        return value

    def show_table_of_all_claims():
        data = SetGetData.fetch_data()
        df = pd.DataFrame(data, columns=['id', 'claim_date', 'project', 'part_name', 'part_id', 'customer', 'customer_claim_id', 'type_of_claim', 'claimed_failure', 'amount_of_claimed'])
        return df

    def export_table_to_excell_cw():
        table = SetGetData.show_table_of_all_claims()
        try:
            today = date.today()
            name_of_file = f'table_of_claims{today}.xlsx'
            table.to_excel(name_of_file)
            print('Table exported to the file: ',name_of_file)
        except Exception as e:
            print('error', e)

    def get_column_names_db(table):
        conn = pg.connect(
            database='Qualita_test',
            user='postgres',
            password='kco177',
            host='localhost',
            port='5432')

        conn.autocommit = True
        cursor = conn.cursor()
        cursor.execute(f'SELECT * FROM {table}')
        column_names = [desc[0] for desc in cursor.description]
        # for i in column_names:
        #    print(i)
        conn.commit()
        conn.close()
        return column_names

    def warhouse_stock(part):
        # get stock for current part
        # use exported excell table from Logistic db
        # need to made dictionary because of differnce in dbases
        project_parts_dictionary = Dic.p8_parts_dictionary #{'AR': 'Antireflection', 'CA': 'Central Airvent', 'CH': 'Cupholder', 'GB': 'Glovebox', 'DP': 'DP upper skin', 'DF':'Deco front'}
        if part.endswith(('LHD', 'RHD')):
            split_string = part.split()
            part_shortcut = split_string[0]
            appendix = split_string[1]
            if part_shortcut in Dic.ROM_parts:
                part_name = project_parts_dictionary[part_shortcut] #+ ' ' +appendix
        else:
            part_name = part
        #print('part name: ',part_name)

        try:
            print('collecting warehouse data from logistic db...')
            # improve path to SQL query
            workbook = load_workbook(filename= Dic.path_warehosue_stock) #'Aktuální stav skladů.xlsx')

            sheet = workbook.active
            data_warehouse_stock = []
            #if part_name in Dic.ROM_parts:
            #    part_name = (project_parts_dictionary[part_name])

            for row in sheet.iter_rows(min_row=3, min_col=1, values_only=True):
                if part_name in row[0]:
                    data_warehouse_stock.append((row[0], row[4]))
            df_wst = pd.DataFrame(data_warehouse_stock)
            print(df_wst)
            return df_wst
        except:
            print('error in logistic data')
class Aggregation:
    def select (claim_id):
        conn, cur = SetGetData.connect_claim()

        try:
            cur.execute(f'SELECT * FROM claims where id = {claim_id}')

        except:
            print('error !')

        # store the results in data
        data = cur.fetchall()
        return data


    def select_decoded (claim_id):
        conn, cur = SetGetData.connect_claim()

        try:
            cur.execute(f'SELECT * FROM decoded WHERE claim_id = {claim_id}')

        except:
            print('error !')

        # store the results in data
        data = cur.fetchall()
        return data






    def update_db_claims(id, part_name, column):

        sql = f'UPDATE claims SET {column} = %s WHERE id = %s'
        conn = None
        update_rows = 0
        try:
            conn, cur = SetGetData.connect_claim()
            cur.execute(sql, (part_name, id))
            update_rows = cur.rowcount
            conn.commit()
            cur.close()
        except Exception as e:
            print(e)
        finally:
            if conn is not None:
                conn.close()
        return update_rows

class Report:
    ''' For making quality reports'''

    # indexes for test db claim table:
    # 0 = 'id',
    # 1 = 'claim_date',
    # 2 = 'project',
    # 3 = 'part_name',
    # 4 = 'part_id',
    # 5 = 'customer',
    # 6 = 'customer_claim_id',
    # 7 = 'type_of_claim',
    # 8 = 'claimed_failure',
    # 9 = 'amount_of_claimed'

    def claim_path(claim_id):
        claim_path = Dic.path_claim_folder #'C:/Users/START/PycharmProjects/databaze/claimfolder'

        claim_data = Aggregation.select(claim_id)
        current_time = datetime.datetime.now()
        IDclaim = claim_id
        folder_claim_name = 'claim' + str(current_time.year) + '_' + str(IDclaim) + '_' + str(claim_data[0][8])
        direction = ((claim_path) + '/' + (folder_claim_name))

        folder_name = 'my_folder'

        if not os.path.exists(direction):
            Report.make_claim_folder(direction)
        else:
            print(f'Folder{folder_claim_name} already exist')

        #print('direction from claim_path method', direction)
        return direction

    def make_claim_folder(direction):

        os.mkdir(direction)
        print('New claim folder saved:', direction)
        return(direction)

    def qa(claim_id, direction):

        #get data and picture, need to improve
        picture_path = 'picture.jpg'
        claim_data = Aggregation.select(claim_id)
        current_time = datetime.datetime.now()
        IDclaim = claim_id

        #set variables:
        name_qa = 'QA_' + str(current_time.year) + '_' + str(IDclaim) + '_' + str(claim_data[0][8]) + '.xlsx'
        box_label = '100% check ' + str(claim_data[0][6])+ ' '+ str(claim_data[0][8])

        #make quality alert
        #fill the template
        try:
            # define position for cells in excell template:
            # customer, project, part_name, date, failure, qa_id, box_label
            qa_position =['C5','R5','G6','R6','A8','D7','M45']
            # define position in outputed tuple from claim table
            # for set position [0][x]
            claim_position = [5,2,3,1,8]

            workbook_qa = load_workbook('QA_template.xlsx')
            sheet = workbook_qa.active


            # add data from list qa_position (excell cells)
            # and claim_position (db claim_table output)
            qa_position_length = len(qa_position) - 2
            for i in range (0, qa_position_length):
                a = str(qa_position[i])
                b = (claim_position[i])
                sheet[a] = claim_data[0][b]

            # add 30+ days qa validity
            today = datetime.date.today()
            qa_validity = today + timedelta(days=30)
            sheet['O7'] = qa_validity

            # add QA_id and box label
            sheet[qa_position[5]] = str(current_time.year) + '_' + str(IDclaim)
            sheet[qa_position[6]] = box_label

            #save the finished qa
            #workbook_qa.save(filename = name_qa)
            p = (direction)+'\\'+(name_qa)
            workbook_qa.save(p)
            print('quality alert',name_qa)

        except Exception as e:
            print('error in report qa!', e)

        #make box label
        #fill template
        try:
            label_position = ['A1', 'A11', 'A21', 'A31', 'A41']
            workbook_label = load_workbook('Label_template.xlsx')
            sheet = workbook_label.active
            label_position_length = len(label_position)
            for i in range(0, label_position_length):
                a = str(label_position[i])
                sheet[a] = box_label
            name_label = 'Box_label'+ str(box_label)+'.xlsx'
            p = (direction) + '\\' + (name_label)
            workbook_label.save(p)
            #workbook_label.save(filename= name_label)
            print(name_label, 'saved')
        except:
            print('error in report label!')


    def _8dreport(id):
        pass

    def make_complete_claim_folder(claim_id):
        directory = Report.claim_path(claim_id)
        Report.qa(claim_id, directory)
        Training_list.train(claim_id, directory)
class Training_list:
    # Trida k vytezovani dat z tydenich prehledu vztvorenych v excellu.
    # Excel musi byt ve formatu *.xlsx.
    # Ver 01/23
    def set_cw_range():
        cw2 = Time.current_cw()
        cw1 = cw2 - 2
        cws = [cw1, cw2]
        return cws

    def recognise_line(project, part):
        #remove from part LHD/RHD
        if 'LHD' or 'RHD' in part:
            part = part.replace(' LHD', '') or part.replace(' RHD','')

        print ('part from recognise line: ', part)

        try:
            ROM_parts = Dic.ROM_parts #('AR', 'CA', 'CH', 'GB', 'DF')
            if part in ROM_parts:
                line_value = 7
            else:
                if project== 'P8' and part == 'IP':
                    line_value = 2
                elif project== 'P8' and 'DP' in part:
                    line_value = 3
                else:
                    print('no data')
            return line_value
        except:
            print('Error or no data for', part)

        # Vstup: (CW_01, CW_02, linka ESA1 = 1; ESA2 = 3; ESA3 = 3; ESA6 = 4; PIP1 = 5; PIP2 = 6, ROM =7)

    def nacti_operatora (cw_01, cw_02, linka):
        # vraci zastoupeni operatoru na linkach po smenach.
        # Vstup: (CW_01, CW_02, linka ESA1 = 1; ESA2 = 3; ESA3 = 3; ESA6 = 4; PIP1 = 5; PIP2 = 6)
        # duplicita CW ve vypisu operator 1 a 2

        path = Dic.path_production_folder #'C:/Users/START/PycharmProjects/databaze/production_folder'
        excel_name =   Dic.name_production_list_SPRAY #'/Průběh směny P8 SPRAY R011 KW'
        i = None
        data = []
        #data01 = []

        # open excel files in given CWs range for operator_01 position
        for i in range(cw_01, cw_02+1):
            if i < 10:
                week = f'{excel_name}0{i}'
                #idweek = i
            else:
                week = f'{excel_name}{i}'
                #idweek = i
            pripona = '.xlsx'
            cw = path + week + pripona

            # find operators in spray excel for spray
            obsazeni_operatora_01 = pd.read_excel(cw, 'Souhrny', skiprows=197, nrows=18, usecols='B:Q')
            data.append(obsazeni_operatora_01.loc[linka])
            #data01.append(idweek)
            print('load data from:', cw)

            # open excel files in given CWs range for operator_02 position
            for i in range(cw_01, cw_02 + 1):
                if i < 10:
                    week = f'{excel_name}0{i}'
             #       idweek = i
                else:
                    week = f'{excel_name}{i}'
              #      idweek = i
        pripona = '.xlsx'
        cw = path + week + pripona
        # find operators in spray excel
        obsazeni_operatora_02 = pd.read_excel(cw, 'Souhrny', skiprows=218, nrows=18, usecols='B:Q')
        obsazeni_operatora_02 = pd.read_excel(cw, 'Souhrny', skiprows=218, nrows=18, usecols='B:Q')
        data.append(obsazeni_operatora_02.loc[linka])
        #data01.append(idweek)
        print('load data 02 from:', cw)

        obsazeni_linky = pd.DataFrame(data)

        obsazeni = pd.DataFrame(obsazeni_linky)
        obsazeni = obsazeni.sort_index()
        obsazeni = obsazeni.rename_axis('CW')

        return obsazeni

    def rom_vypis_operatoru(cw_01):
        path = 'C:/Users/START/PycharmProjects/databaze/production_folder'
        excel_name = '/Průběh směny P8 B8 ROM R12 KW'
        i = None
        data = []

        # open excel files in given CWs range for operator_01 position
        if cw_01 < 10:
            week = f'{excel_name}0{cw_01}'
        else:
            week = f'{excel_name}{cw_01}'
        pripona = '.xlsx'
        cw = path + week + pripona
        obsazeni_operatora_01 = pd.read_excel(cw, 'OBSAZENÍ', skiprows=1, nrows=150, usecols='A:G')
        print('load data for training from:', cw)
        df = pd.DataFrame(obsazeni_operatora_01)

        #make one col from excell cols
        single_col = pd.concat([df[col] for col in df.columns])
        single_col.dropna(inplace=True)
        #prepare the df to unify training list method
        seznam_operatoru = pd.DataFrame(single_col)
        vypis_operatoru_beta = seznam_operatoru.reset_index()  # zmen nazev indexu z defaultniho na 'jmeno'
        vypis_operatoru = vypis_operatoru_beta.rename(columns={'index': 'x', '0': 'index'})
        vypis_operatoru = vypis_operatoru.rename(columns={0: 'index'})
        vypis_operatoru = vypis_operatoru.drop('x', axis=1)
        return vypis_operatoru

        # uloz do template / je stejny pro vsechny
        workbook_training_list = load_workbook('training_list_template.xlsx')
        sheet = workbook_training_list.active
        # add list of operators into right cells in excell list of operators in the column 'index'
        df = vypis_operatoru
        start_cell = sheet.cell(row=14, column=2)
        for i, value in enumerate(df['index']):
            cell = start_cell.offset(row=i, column=0)
            cell.value = value

        workbook_training_list.save('training_list_test.xlsx')

    def vypis_operatory(cw_01, cw_02, linka):

        # vraci soupis operatoru na linkach a jejich datum nastupu. Vstup: (CW_01, CW_02, linka ESA1 = 1; ESA2 = 3; ESA3 = 3; ESA6 = 4)
        # porovnava s daty kmenovych a agenturnich zamestnancu
        #### POZOR - aktualizuje se 1* mesicne - muze menit format

        dat = Training_list.nacti_operatora(cw_01, cw_02, linka)

        # pro cw vytvor soupis operatoru
        seznam_operatoru = dat.melt()
        vypis_operatoru_beta = pd.DataFrame(seznam_operatoru['value'].value_counts())
        vypis_operatoru_beta = vypis_operatoru_beta.reset_index()  # zmen nazev indexu z defaultniho na 'jmeno'
        vypis_operatoru = vypis_operatoru_beta.iloc[3:]  # odstran prvni tri radky

        # pokud P8 IP iteruj na linku 2
        return (vypis_operatoru)

    def training_list(project,part,cw,claim_ID,direction):
        #default cw = 0
        current_time = datetime.datetime.now()
        claim_data = Aggregation.select(claim_ID)
        name_tl = (f'trainining_list_{current_time.year}_ {claim_ID}_{claim_data[0][8]}.xlsx')

        if cw == 0:
            cws = Training_list.set_cw_range()
            a = cws[0]
            b = cws[1]
        else:
            #pokud je zadan konkretni kalendarni tyden
            #rozptyl tydnu je +/- 1 kalendarni tyden
            a = cw - 1
            b = cw + 1

        c = Training_list.recognise_line(project, part)
        if c == 2:
            df = Training_list.vypis_operatory(a,b,c)
        elif c == 7 or 3:
            df = Training_list.rom_vypis_operatoru(b)
        else:
            print('production line does not exist')

        try:
            #open workbook template
            workbook_training_list = load_workbook(Dic.training_list_template) #'training_list_template.xlsx')
            sheet = workbook_training_list.active

            #add list of operators into right cells in excell list of operators in the column 'index'
            start_cell = sheet.cell(row = 14, column = 2)
            for i, value in enumerate(df['index']):
                cell = start_cell.offset(row=i, column=0)
                cell.value = value

            #add values from claims
            #training, failure
            training_list_position = ['C5', 'A8']
            # define position in outputed tuple from claim table
            # for set position [0][x]
            #claim_position = [8]

            trainig_list_length = len(training_list_position)
            for i in range(0, trainig_list_length):
                a = str(training_list_position[i])
                sheet[a] = (f'trainining_list_{current_time.year}_ {claim_ID}_{claim_data[0][8]}')


            #name_training_list = 'training_list' + str(current_time.year) + '_' + str(IDclaim) + '_' + str(claim_data[0][8] + '.xlsx'
            #save finished trining list
            #workbook_training_list.save('trainig_list.xlsx')
            p = (direction) + '\\' + (name_tl)
            workbook_training_list.save(p)
            #_name = 'claim' + str(current_time.year) + '_' + str(IDclaim) + '_' + str(claim_data[0][8] + '.xlsx')
            #direction = ((claim_path) + '/' + (claim_name))

            print(f'training_list{name_tl}saved')

        except:
            print('error in training list')

    def train (claim_Id, direction):
        try:
            data = Aggregation.select(claim_Id)
            # define position in outputed tuple from claim table
            # for set position [0][x]
            project = data[0][2]
            part = data[0][3]
            cw_default = 0 #default


            Training_list.training_list(project,part,cw_default,claim_Id,direction)
            print('training_list(project, part, cw, claim_ID, direction):')
        except Exception as e:
            print(e)
#driver functions
#SetGetData.insert_claim('2023-03-09', 'P8', 'DP', 'a', 'SMRC', 'b8mmjfo', 'A', 'flek', 20, 320)
#Report.make_complete_claim_folder(19)
#Report.prepare_qa_mail_message(19)

#id = 6
#Report.qa(id, directory)
#Aggregation.update(5)
#direction = 'C:/Users/START/PycharmProjects/databaze/claimfolder'
#Training_list.train(5, direction)


###>>> update na zaklade claim id (id, new_value, column)
#print(Aggregation.select(4))
#Aggregation.update_db_claims(4,'D0323116', 'part_id')
#print(Aggregation.select(4))
