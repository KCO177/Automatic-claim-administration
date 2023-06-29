from db_quality import SetGetData, Time, Aggregation
from outlook_management import Mail_output
import pandas as pd
import matplotlib.pyplot as plt
import os
import datetime
from datetime import timedelta
from openpyxl import load_workbook
from qualita_db_dictionary import MyDict as Dic

class Db_data:

    def collect_data_subject(subject):
        #collecting and 'filter' data from existing methods based on click on the mail in list
        nc = Mail_output.text_output(subject)  # new claim from e-mail according clicked subject
        #[date, project, part, part_id, customer, claim_number, 'claim', failure, nok_amount, stock_amount]
        # nc [1] = project, nc[2] = part, nc[7] = nok_amount, nc[9] = stock amount

        #get values for variables from mail output
        project = nc[1]
        part = nc[2]
        nok_amount = nc[7]
        cust_stock_amount = nc[9]

        #get data from warehouse stock_dbase
        if project == 'P8':
            stock = SetGetData.warhouse_stock(part)
        else:
            stock = str('No Data for project', project)

        Mail_data_collected = [project, part, nok_amount, cust_stock_amount, stock]

        return Mail_data_collected

    def collect_data_claim_id(claim_id):
        data = Aggregation.select(claim_id)
        print(data)
        project = data[0][2]
        part = data[0][3]
        nok_amount = data[0][9]
        cust_stock_amount = data[0][10]

        if project == 'P8':
            stock = SetGetData.warhouse_stock(part)
            print('collect_data_claim_id', stock)
            if not stock.empty:
                sum_stock = stock[1].sum()
            else:
                sum_stock = 0
        else:
            sum_stock = 0

        Mail_data_collected = [project, part, nok_amount, cust_stock_amount, sum_stock]
        #print(Mail_data_collected, 'Mail_data_collected')
        return Mail_data_collected


    def insert_log_fin(claim_id, customer_stock, warehouse_stock, sorting_cost_assumption, claim_cost_assumption):

        conn, cur = SetGetData.connect_claim()
        try:
            cur.execute(
                'INSERT INTO logistic_financial VALUES (%s, %s, %s, %s, %s)',
                (claim_id, customer_stock, warehouse_stock, sorting_cost_assumption, claim_cost_assumption))

            print('new financial assumption saved for claim', claim_id)
        except Exception as e:
            print('error', e)

        # comiting the transaction
        conn.commit()

class Expedition_df:
    def create_df(cw, part, sorting_takt, hour_price):

        # >>> path
        files = os.listdir('C:/Users/START/PycharmProjects/databaze/log_data')
        file_name_end = str(cw) + '.xlsx'

        # najdi aktualni expedicni plan
        for file in files:
            if file.endswith(file_name_end):
                excel_file = file
                print('collecting expedition data from file = ', file)

        path = 'C:/Users/START/PycharmProjects/databaze/log_data/' + excel_file

        workbook = load_workbook(filename=path)
        sheet = workbook.active
        sheets = workbook.sheetnames
        current_sheet = 'Expedition CW' + str(cw)
        ws = workbook[current_sheet]

        # najdi posledni vyplneny radek
        index = 0
        for cell in ws['J']:
            if cell.value != None:
                index = index + 1
            else:
                break
        row_limit = index

        # vytvor dataframe
        columns = ['Date', 'ETA Ascorium', 'Qty', 'Description', 'DN n°']
        df = pd.read_excel(path, sheet_name=current_sheet, usecols=columns, nrows=row_limit)

        # vypln mezery mezi datumy a casy
        df['Date'] = df['Date'].fillna(method='ffill')
        df['ETA Ascorium'] = df['ETA Ascorium'].fillna(method='ffill')

        # preved datum a cas do jednoho dsloupce ve formatu datetime
        df['DateTime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['ETA Ascorium'].astype(str))
        df.drop(columns=['ETA Ascorium', 'Date'], inplace=True)
        column_order = ['DateTime', 'Description', 'Qty', 'DN n°']
        df = df[column_order]

        # filtrovani v df
        now = datetime.datetime.now()
        format_time = now.strftime("%H:%M:%S")
        format_date = now.strftime('%Y-%m-%d')
        #print('PART', part)
        part_name = part #str(Dic.part_id[part])
        print('PART NAME', part_name)
        vyvezeno_df = df.loc[df['DateTime'] < now]
        print('vyvezeno_df', vyvezeno_df)
        part_vyvezeno = vyvezeno_df.loc[vyvezeno_df['Description'].apply(lambda x: x.startswith(part_name))]
        print('part_vyvezeno', part_vyvezeno)
        na_ceste = now - timedelta(hours=8)
        na_ceste_df = part_vyvezeno.loc[part_vyvezeno['DateTime'] > na_ceste]
        na_ceste_df['Status'] = 'on_the_way'
        print('na_ceste_df', na_ceste_df)

        pripraveno_df = df.loc[df['DateTime'] > now]
        print('PART NAME pripraveno', part_name)
        print('pripraveno_df', pripraveno_df)
        pripraveno_df.loc[:, 'Description'] = pripraveno_df['Description'].astype(str)
        part_pripraveno_df = pripraveno_df.loc[pripraveno_df['Description'].apply(lambda x: x.startswith(part_name))]
        #part_pripraveno_df['Status'] = 'planned'
        print('part_pripraveno_df', part_pripraveno_df)

        df = pd.concat([na_ceste_df, part_pripraveno_df])
        #print(df)

        sorting_cost = []
        admin_cost = 1 * hour_price

        for num_of_parts in df['Qty']:
            if num_of_parts > 0:
                sorting_assumption = round((num_of_parts / sorting_takt * hour_price), 2)
            else:
                sorting_assumption = 0
            sorting_cost.append(sorting_assumption)

        df['Sorting_assumption €'] = sorting_cost
        df['Cum_sum €'] = df['Sorting_assumption €'].cumsum()

        #print(part, now)
        #print(df)
        return df

    def exped_sort_assume(expedition_df):
        now = datetime.datetime.now()
        plus_one_day = now + timedelta(hours=24)
        assumption_delta = plus_one_day.strftime('%Y-%m-%d')
        assumption_df = expedition_df.loc[expedition_df['DateTime'] < assumption_delta ]
        suma_euro = round(assumption_df['Sorting_assumption €'].sum(),2)
        #print(assumption_df)
        print('on the way + 24 hours deliveries sorting assumption: ', suma_euro)
        return suma_euro

    def next_exped_df(expedition_df):
        print('next_exped_df', expedition_df)
        #Expedition_df.plot_sorting_assumption(expedition_df)
        now = datetime.datetime.now()
        plus_one_day = now + timedelta(hours=24)
        assumption_delta = plus_one_day.strftime('%Y-%m-%d')
        exped_df = expedition_df.loc[expedition_df['DateTime'] < assumption_delta]
        print('next_exped_df_exped_df', exped_df)
        return exped_df

    def plot_sorting_assumption(expediton_df):
        df = expediton_df
        plt.plot(df['DateTime'], df['Cum_sum €'])
        plt.xlabel('Date')
        plt.ylabel('sorting_assumption €')
        plt.title('sorting assumption cumulative €')
        plt.xticks(rotation=90)
        plt.show()

    def next_exped(claim_id):  #>>> claim ID
        cw = Time.current_cw()
        claim_data = Db_data.collect_data_claim_id(claim_id)  # [project, part, nok_amount, cust_stock_amount, stock]
        warehouse_stock_assumption = claim_data[3]
        part = claim_data[1]
        # condition for translate between different names of parts using in excells
        if part == 'IP':
            part = 'L'
        #(cw, part, sorting_takt, hour_price)
        input_df = Expedition_df.create_df(cw, part, sorting_takt=180, hour_price=38)
        print('input_df', input_df)
        exped_summary = Expedition_df.next_exped_df(input_df)
        print('exped_summary', exped_summary)
        exped_summary.drop(columns=['Sorting_assumption €', 'Cum_sum €'], inplace=True)

        #print(exped_summary)

        return exped_summary

class Financial:
    def sorting_assumption(claim_id):
        #collect claim data
        cw = Time.current_cw()
        #subject = 'claim cupholder'
        claim_data = Db_data.collect_data_claim_id(claim_id) #[project, part, nok_amount, cust_stock_amount, stock]
        warehouse_stock_assumption = claim_data[3]
        part = claim_data[1]
        #condition for tranaslate between different names of parts using in excells
        if part == 'IP':
            part = 'L'

        exped_assumption = Expedition_df.exped_sort_assume(Expedition_df.create_df(cw, part,sorting_takt=180, hour_price=38))
        cost_assumption = warehouse_stock_assumption + exped_assumption

        return cost_assumption
        print('total assumption for sorting is: ', Financial.sorting_assumption(), '€')

    def insert_new_item(claim_id):
        data = Db_data.collect_data_claim_id(claim_id)
        print(data)
        customer_stock = int(data[3])
        warehouse_stock = int(data[4])
        sorting_cost = Financial.sorting_assumption(claim_id)
        claim_cost_assumption = 250
        Db_data.insert_log_fin(claim_id, customer_stock, warehouse_stock, sorting_cost, claim_cost_assumption)

#Driver functions:
#Expedition_df.plot_sorting_assumption()
#df = Expedition_df.create_df(22,'L',180,38)
#Expedition_df.sort_assume(df)
#Financial.sorting_assumption(subject) #>>> vraci sorting assumption
#subject = 'claim cupholder'
#Expedition_df.next_exped(subject)
