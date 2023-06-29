import win32com.client
from db_quality import SetGetData, Aggregation, Report
from qualita_db_dictionary import MyDict as Dic
from datetime import datetime
from outlook_management import Mail_text
import datetime
from qualita_db_dictionary import MyDict as Dic
import cv2
import pytesseract
import pandas as pd
from collections import ChainMap
from pyzbar.pyzbar import decode as qr_decode
from pylibdmtx.pylibdmtx import decode as matrix_decode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import openpyxl
import os
import re
from PIL import Image


#>>>claimfolderpath = Report.claim_path(claim_id)
#claimfolder uz musi bzt vytvoreny


class Mail_attachment:
    #>>> path to the file

    def save_attachment(path_to_claim_file, subject):
        sub = str(subject)
        outlook = win32com.client.Dispatch('outlook.application').GetNamespace('MAPI')

        inbox = outlook.GetDefaultFolder(6)
        messages = inbox.Items  # .Folders
        #current_time = datetime.datetime.now()
        #received_dt = current_time - timedelta(days=7)
        #received_dt = received_dt.strftime('%m/%d/%Y %H:%M %p')
        #messages = messages.Restrict("[ReceivedTime] >= '" + received_dt + "'")
        #messages = messages.Restrict("[SenderEmailAddress] = 'petr.koscelnik@ascorium.com'")
        messages = messages.Restrict(f"[subject] = {sub}")  #
        message = messages.GetFirst()  # GetLast(), GetNext(), . Attachments

        # message.Subject #.Body, .To, .Recipients, .Sender, .Sender.Address
        #message_text = message.subject + ' ' + message.body

        outputDir = path_to_claim_file
        for message in messages:
            attachments = message.attachments
            s = message.sender
            for attachment in attachments:
                print(f'attachment {attachment.FileName}')
                attachment.SaveASFile(os.path.join(outputDir, attachment.FileName))
                print(f'attachment {attachment.FileName} from {s} saved')

    def get_list_of_file(path_to_claim_file):
        list_of_files = os.listdir(path_to_claim_file)
        return list_of_files
        #for file in files:
        #    print(file)
class Picture_resizing_for_qa:

    def cell_size(excel_file= Dic.quality_alert_template):
        #>>> improve cell coordinate to argument
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        cell_height = 2.22 * worksheet.row_dimensions[4].height
        cell_width = 4.25 * worksheet.column_dimensions['A'].width

        #print(f"The width of cell A4 is {cell_width} pixels.")
        #print(f"The height of cell A4 is {cell_height} pixels.")

        return (cell_width, cell_height)
    def picture_list(path_to_the_file):
        list_of_files = Mail_attachment.get_list_of_file(path_to_the_file)
        suffix = ['.jpg', '.png', 'bmp', '.jpeg', '.tiff']
        filtered_list = [item for item in list_of_files if any(item.endswith(word) for word in suffix)]
        print('in attachment are ready for analysing', filtered_list)
        return filtered_list
    def add_picture_in_QA(path_to_the_resize_folder, path_to_existing_qa, size_list):
        # >>> finish save path
        # >>> solve anchors
        #print('path_to_existing_qa from add_picture_in_QA method> ', path_to_existing_qa)
        workbook_qa = load_workbook(path_to_existing_qa) #'C:/Users/START/PycharmProjects/databaze/QA_template.xlsx')
        sheet = workbook_qa.active

        list_of_pictures = Picture_resizing_for_qa.picture_list(path_to_the_resize_folder)
        path_to_folder = path_to_the_resize_folder.replace('\\', '/')
        #print(size_list)

        # index = size_list
        cell_dimension = Picture_resizing_for_qa.cell_size(path_to_existing_qa)
        cell_height = cell_dimension[0]

        # posun radku
        row_shift = []
        row_shift_nominal = []

        cumulative = 0
        for size in size_list:
            cell_anchor = round(size / cell_height)
            cumulative += cell_anchor
            row_shift.append(cumulative)
            row_shift_nominal.append(cell_anchor)

        #print('posun radku', row_shift)
        #print('nominal', row_shift_nominal)

        anchor_cell_list = []

        for i in list_of_pictures:

            shift_c = 0
            # >>> placemenet + index / const_row_pix
            # >>> podminka pokud je posledni picture presun se do druheho sloupce
            picture_list_index = list_of_pictures.index(i)
            #print('picture_list_index: ', picture_list_index)

            # neplati pro prvni obrazek, zacina na radku 13
            row = 13

            if picture_list_index > 0:
                row = str(13 + row_shift[picture_list_index - 1])

            range_row = 12 + row_shift[picture_list_index]

            #print('spodni limit', range_row)
            if range_row <= 43:
                anchor = 'A' + str(row)
                anchor_cell_list.append(anchor)

            else:
                # pokud presahne rozsah obrazku radek 43 posun obrazek do sloupce K

                # a = int(row)
                shift_a = row_shift[picture_list_index]
                # pokud neni posledni index v listu
                if row_shift.index(shift_a) + 1 < len(row_shift):
                    shift_b = row_shift[picture_list_index + 1]
                    shift = (shift_b - shift_a)
                else:
                    shift = row_shift.index(shift_a) - 1

                # pokud neni prvni index v listu vrat posledni anchor
                #print(anchor_cell_list)
                #print(picture_list_index)
                #print(anchor_cell_list[int(picture_list_index) - 1])
                previous_anchor = anchor_cell_list[int(picture_list_index) - 1]
                #print('previous anchor', previous_anchor)

                # pokud posledni anchor je v prvnim sloupci -> pravidlo pro prvni obrazek
                if previous_anchor.startswith('A'):
                    # print('row shift: ',row_shift)

                    # print('posun', shift)
                    # print(row_shift)
                    # print('a',row)
                    row = ((row_shift[picture_list_index - 1]) - (row_shift[picture_list_index - 1]) + 13)
                    anchor = 'J' + str(row)
                    anchor_cell_list.append(anchor)

                else:
                    # neni to prvni obrazek pricti posun
                    # print(row_shift)
                    shift = row_shift_nominal[picture_list_index]
                    # print('posun', shift)

                    # spocitej hodnotu posledniho row => dostan 0 + 13 + shift -1 indexu
                    # >>> previous_row = (row_shift_nominal[picture_list_index-1]) + 13

                    if previous_anchor.startswith('A'):
                        previous_shift = 0
                    else:
                        previous_shift = (
                        row_shift_nominal[picture_list_index - 1])  # - (row_shift[picture_list_index-])
                    # pokud posledni row_shift[picture_list_index -2] startswtith['A'] -> posledni_shift = 0
                    # else: posledni shift =  row_shift[picture_list_index -1] - row_shift[picture_list_index -1]
                    previous_row = int(re.search(r'\d+', previous_anchor).group())

                    # anchor -1 index -> vyjmout int = previous row

                    # print('previous row:', previous_row)
                    row = previous_row + previous_shift
                    # print('row z posledni podminky', row)
                    # next_row = int(row)+int(shift)
                    # print('novy radek', next_row)

                    anchor = 'J' + str(row)
                    anchor_cell_list.append(anchor)

            #print(anchor)
            #print(row_shift_nominal)
            path_to_picture = str(path_to_folder + '/' + i)
            #print('Adding picture: ', path_to_picture, 'into the position', anchor)

            fig = openpyxl.drawing.image.Image(path_to_picture)
            sheet.add_image(fig, anchor)  # 'B14')

            print(os.getcwd())

        ###>>> path to the current QA
        workbook_qa.save(path_to_existing_qa)
    def meassuring_area(excel_file, area):
        workbook = load_workbook(excel_file)

        cell_range = 'A14:T43' #area
        # Extracting individual components

        a = cell_range[0]  # 'A'
        b = cell_range[4]  # 'T'

        def pismena_na_cislo(pismena):
            cislo = 0
            mocnina = 1
            for pismeno in reversed(pismena):
                cislo += (ord(pismeno.upper()) - ord('A') + 1) * mocnina
                mocnina *= 26
            return cislo

        rozsah = 'A:T'
        pocatecni_sloupec, koncovy_sloupec = rozsah.split(':')

        cislo_pocatecniho_sloupce = pismena_na_cislo(pocatecni_sloupec)
        cislo_koncoveho_sloupce = pismena_na_cislo(koncovy_sloupec)

        pocet_sloupcu = cislo_koncoveho_sloupce - cislo_pocatecniho_sloupce + 1




        c = int(cell_range[1:3])  # 14
        d = int(cell_range[5:])  # 29
        diff_row = 31 #d - c #31
        diff_col = pocet_sloupcu #ord(b) - ord(a) + 1  # 20 (using ASCII values)
        #print('diff_col', diff_col, ' = 20')
        #print('diff_row', diff_row, ' = 31')

        worksheet = workbook.active
        cell = Picture_resizing_for_qa.cell_size()
        cell_width = cell[1]
        cell_height = cell[0]
        #cell_height = 2.22 * worksheet.row_dimensions[4].height
        #cell_width = 4.25 * worksheet.column_dimensions['A'].width
        #print('one cell: ', cell_width, cell_height)

        area_height = diff_row * cell_height
        area_width = diff_col * cell_width

        # Print the results
        # print(f"The height of cell A1 is {cell_height} pixels.")
        # print(f"The width of cell A1 is {cell_width} pixels.")

        #print(f"area_width {area_width} pixels.")
        #print(f"area_heigth {area_height} pixels.")

        return (area_width, area_height)
    def area_for_pictures(list_of_pictures):
        # velikost oblasti v pixlech

        area = (Picture_resizing_for_qa.meassuring_area('QA_template.xlsx', 'A13:T43'))
        filtered_list = list_of_pictures  # Mail_attachment.get_list_of_file(path_to_claim_folder)
        number_of_pictures = (len(filtered_list))
        # print(number_of_pictures)

        # kolik je obrazku sud/lich
        if number_of_pictures % 2 == 0:
            heigth_number = number_of_pictures / 2
        else:
            heigth_number = (number_of_pictures + 1) / 2

        # print('in attachment are ready for analysing', filtered_list, 'number of columns with pictures', heigth_number, 'number of rows with picture', 'width_number')

        # pevne stanovene hodnoty
        picture_width = int(area[0] / 2)
        picture_heigth = int(area[1] / (heigth_number))
        #print('from area_for_picture : picture_width', picture_width, 'picture_heigth',picture_heigth)
        return (picture_width, picture_heigth)
    def picture_size(list_of_parts, path_to_the_file):
        size = []
        #dpi = 96
        for picture in list_of_parts:
            path = path_to_the_file + '/' + picture
            image = Image.open(path)
            width, heigth = image.size
            # heigth_in_points = heigth * 72 / dpi
            size.append(heigth)  # _in_points)
            #print('size',size)
        #print('heigths' , size)

        return size
    def picture_resize(path_to_pic, original_shape, new_shape):
        img = Image.open(path_to_pic)
        if img.mode in ("RGBA", "P"): img = img.convert("RGB")

        # resize with aspect ratio
        o_w = original_shape[0]
        o_h = original_shape[1]
        required_w = new_shape[0]
        required_h = new_shape[1]

        #print ('original_w_h: ',o_w, o_h, 'required w_h: ',required_w, required_h)

        #o_w = img.size[0]
        #o_h = img.size[1]

        #print (o_w, o_h)

        # if \
        # o_w >= o_h:
        a = o_w
        b = o_h
        c = required_h
        '''
        else:
            a = o_w
            b = o_h
            c = required_w
        '''
        wpercent = (c / float(a))
        hsize = int((float(b) * float(wpercent)))
        img = img.resize((c, hsize), Image.LANCZOS)

        path = path_to_pic
        a, b = path.rsplit('/', 1)
        b = b.split('.')[0]

        a_w = img.size[0]
        a_h = img.size[1]

        #print ('after resize',a_h, a_w)

        picture_save = a + '/resized_images/' + b + '_resized.jpg'

        img.save(picture_save)
    def open_resize_folder_if_exist(path_to_claim_folder):
        folder_name = path_to_claim_folder + '/' + 'resized_images'

        if not os.path.exists(folder_name):
            os.makedirs(folder_name)
            print(f'Folder {folder_name} created')
        else:
            print(f'Folder{folder_name} already exist')

        return folder_name
    def resize_in_folder(path_to_claim_folder):
        # path_to_claim_folder = 'C:\\Users\\START\\PycharmProjects\\databaze\\claimfolder\\claim2023_42_Bubble'
        path_to_claim_file = path_to_claim_folder.replace('\\', '/')
        # print(path_to_claim_folder)
        path_to_the_resize_folder = Picture_resizing_for_qa.open_resize_folder_if_exist(path_to_claim_folder)
        list_of_pictures = Picture_resizing_for_qa.picture_list(path_to_claim_folder)
        # print(list_of_pictures)
        new_shape = Picture_resizing_for_qa.area_for_pictures(list_of_pictures)
        # print(new_shape)
        for i in list_of_pictures:
            path_to_pic = str((path_to_claim_file) + '/' + (i))
            image = Image.open(path_to_pic)
            width, heigth = image.size

            # heigth_in_points = heigth * 72 / dpi
            #size.append(heigth)  # _in_points)

            original_shape = (width,heigth)

            #Picture_resizing_for_qa.picture_size(path_to_pic)
            Picture_resizing_for_qa.picture_resize((path_to_pic), (original_shape), (new_shape))
        print('path_to_the_resize_folder z resize_in_folder metody', path_to_the_resize_folder)
        return str(path_to_the_resize_folder)
    def find_existing_qa(path_to_the_claim_folder):
        files = os.listdir(path_to_the_claim_folder)
        for file in files:
            qa = (file.startswith('QA'))
            if qa == True:
                break
        return file
    def add_pictures_to_QA (path_to_the_claim_folder):
        Picture_resizing_for_qa.open_resize_folder_if_exist(path_to_the_claim_folder)
        path_to_the_resized_file = Picture_resizing_for_qa.resize_in_folder(path_to_the_claim_folder)
        #print(path_to_the_resized_file)
        existing_qa = Picture_resizing_for_qa.find_existing_qa(path_to_the_claim_folder)
        path_to_the_existing_QA =  path_to_the_claim_folder +'/'+ existing_qa

        # 'C:/Users/START/PycharmProjects/databaze/claimfolder/claim2023_29_Bubble/resized_images'
        #path_to_the_template = 'C:/Users/START/PycharmProjects/databaze/QA_template.xlsx'
        list_of_pictures = Picture_resizing_for_qa.picture_list(path_to_the_resized_file)
        Picture_resizing_for_qa.area_for_pictures(list_of_pictures)
        picture_width = Picture_resizing_for_qa.picture_size(list_of_pictures, path_to_the_resized_file)
        #one_cell_size = Picture_resizing_for_qa.cell_size(path_to_the_template, 'A4')
        Picture_resizing_for_qa.add_picture_in_QA(path_to_the_resized_file, path_to_the_existing_QA, picture_width)
class Picture_processing:

    def ocr_from_label(picture, path_to_claim_file):  # (path)
        # inputed image path >>> replace forclaim folder
        #image_path = 'C:/Users/START/PycharmProjects/databaze/mail_test/'+ (picture)
        image_path = (path_to_claim_file)+'/'+(picture)

        try:
            pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            # Open the image file using Pillow
            with Image.open(image_path) as img:
                # Convert the image to grayscale
                img = img.convert('L')

                # Use pytesseract to extract the text from the image
                text = pytesseract.image_to_string(img)
            #print(text)
            return text
        except Exception as e:
            print(e)
    def data_matrix_from_label(picture, path_to_claim_file):
        print('decoding....',picture)
        image_path = (path_to_claim_file)+'/'+(picture)
        # Read file using OpenCV
        try:
            code = matrix_decode(cv2.imread(image_path))
            decoded_code = code[0]
            text_in_bytes = decoded_code[0]
            text = text_in_bytes.decode('utf-8')
            print(type(text))
            print('in',picture,'decoded', text)
        except:
            text = 'no readable data on image'
        return text
    def qr_bar_from_label(picture, path_to_claim_file):
        image_path = (path_to_claim_file) + '/' + (picture)
        print('decoding...',picture)
        # Read file using OpenCV
        try:
            decode =(qr_decode(cv2.imread(image_path))) #using of pyzbar decoder usable also for bar codes
            #print('kod: ',decode)
            #qcd = cv2.QRCodeDetector() #work only with QR
            #img = cv2.imread(image_path) #work only with QR
            #code = qcd.detectAndDecodeMulti(img) #work only with QR
            data = decode[0]
            text_in_bytes = data[0]
            text = text_in_bytes.decode('utf-8')
            print('in', picture,'decoded', text)
        except:
            text = 'no readable data on image'
        return text
    def get_text_from_attachment(path_to_claim_file, method):
        list_of_files = Mail_attachment.get_list_of_file(path_to_claim_file)

        #print(list_of_files)
        suffix =['.jpg','.png','bmp','.jpeg','.tiff']
        filtered_list = [item for item in list_of_files if any(item.endswith(word) for word in suffix)]
        print('in attachment are ready for analysing', filtered_list)
        text = []

        for picture in filtered_list:
            try:
                if method == 'OCR':
                    word = (Picture_processing.ocr_from_label(picture, path_to_claim_file))
                elif method == 'data_matrix':
                    word = (Picture_processing.data_matrix_from_label(picture, path_to_claim_file))
                elif method == 'QR_Bar':
                    word = (Picture_processing.qr_bar_from_label(picture, path_to_claim_file))
                #elif method == 'all'

                text.append(word)
            except:
                text = 'no readable data on picture'
        return text
class Read_from_attachment:

#Read_from_attachment.update_claim_from_attachment()
    def decode_from_code(decoded_list):
        #print(decoded_list)
        # D03231162306211511828211359
        # rozpad kódu D0323116 23 062 11:51:18 28 2 1 1359
        # part_id  Y   D  HH:MM:SS Linka WE cav denní sekvenční číslo

        decoded_codes = []
        #print(decoded_codes)
        for item in decoded_list:
            if item == 'no readable data on image' or None:

                tuple_item = None
            else:
                # Extract the different parts of the item
                part_id = item[:8]
                prod_day = int(item[8:10])
                prod_month = int(item[10:13])
                prod_time = f"{item[13:15]}:{item[15:17]}:{item[17:19]}"
                prod_line = int(item[19:21])
                prod_we = int(item[21:22])
                prod_kav = int(item[22:23])
                prod_seq_id = int(item[23:])

                # Format the parts into a tuple and append to list2
                tuple_item = (part_id, prod_day, prod_month, prod_time, prod_line, prod_we, prod_kav, prod_seq_id)
            decoded_codes.append(tuple_item)
        #print('decoded_codes_method', decoded_codes)
        return (decoded_codes)
    def update_claim_from_attachment (claim_id, path_to_claim_file, method):
        #>>> POUZE PRO: PART, PROJECT, PART_id,  date of production
        #priprava var pro analyzu textu:

        message = Read_from_attachment.decode_from_code(Picture_processing.get_text_from_attachment(path_to_claim_file, method))

        print('decoded message', message)
        #inv = Dic.invert_dictionary(Dic.p8_parts_dictionary)
        #book = ChainMap(Dic.part_id, Dic.p8_parts_dictionary, inv, Dic.project_dictionary, Dic.projects, Dic.customers)
        # fuzzy_tuple = Mail_text.fuzzy_text(message,book,100)
        # print('Finds in pictures: ', fuzzy_tuple)
        customer_list = []
        project_list = []
        times = []
        dates = []
        parts = []
        parts_id = []
        for item in message:
            if item != None:
                print(item)

                # hledani part a part_ID
                part_id = (Mail_text.find_part_id(item, part=None))
                part = (Mail_text.find_part(part_id))
                parts.append(part)
                parts_id.append(part_id)

                # hledani projektu
                project = Mail_text.find_project(message)
                if project == None and part != None:
                    project_fuzz = Mail_text.fuzzy_text(part, Dic.project_dictionary, 0)
                    project = Dic.project_dictionary[project_fuzz[0]]
                else:
                    project = None
                project_list.append(project)

                # hledani customer
                dictionary = Dic.invert_dictionary(Dic.customers)
                if project != None:
                    if project in dictionary:
                        customer = (dictionary[project])

                else:
                    customer = None
                customer_list.append(customer)

                # stanoveni data a casu
                time = item[3]
                print('time: ', time)
                day_num = str(item[2])
                year = str(item[1] + 2000)
                day_num.rjust(3 + len(day_num), '0')
                res = datetime.datetime.strptime(year + "-" + day_num, "%Y-%j").strftime("%m-%d-%Y")
                dates.append(res)
                times.append(time)

        # tvorba dataframu z listu
        data = {'claim_id': claim_id,
                'project': project_list,
                'customer': customer_list,
                'part_id': parts_id,
                'part': parts,
                'date': dates,
                'time': times
                }
        df = pd.DataFrame(data)
        print(df)

        #>>> if not empty
        # Projdi každý řádek DataFrame a uloz do db
        for index, row in df.iterrows():
            hodnoty_radku = list(row.values)
            print('insert data:', hodnoty_radku, 'in dbase')
            SetGetData.insert_decoded(hodnoty_radku)

        return(df)
    def decoded_unique_values(claim_id):
        # najdi unikatni hodnoty
        input_list = Aggregation.select_decoded(claim_id)

        unikatni_hodnoty = {}
        # Projdi každý řádek vstupního listu
        for row in input_list:
            for index, value in enumerate(row):
                # Pokud je hodnota na daném indexu již v unikatních hodnotách
                if index in unikatni_hodnoty:
                    # Porovnej hodnoty
                    if unikatni_hodnoty[index] != value:
                        # Pokud se liší, spoj je pomocí oddělovače ','
                        unikatni_hodnoty[index] = f"{unikatni_hodnoty[index]}, {value}"
                else:
                    # Pokud hodnota na daném indexu ještě není v unikatních hodnotách, přidej ji tam
                    unikatni_hodnoty[index] = value

        # Vytvoř výsledný list z unikátních hodnot
        result_list = list(unikatni_hodnoty.values())

        # Výpis výsledného listu
        return result_list
    def get_datalist_for_update(claim_id):
        result_list = Read_from_attachment.decoded_unique_values(claim_id)
        unique_values = result_list[2:6]
        print('unique_values before', unique_values)
        unique_values[1], unique_values[2] , unique_values[3]= unique_values[2], unique_values[3],unique_values[1]
        return unique_values
    def update_from_img_to_db(claim_id):
        #RUN metoda pro import dat z obrazku do db
        path_claim_file = str(Report.claim_path(claim_id))
        analyzed_codes_from_picture = Read_from_attachment.update_claim_from_attachment(claim_id, path_claim_file,'data_matrix')
        #df = analyzed_codes_from_picture
        #print(df)
    ###>>> update claim table in db
    def update_claim_table_from_db(claim_id):
        #RUN metoda pro update claim tabulky z decode tabulky
        #>>> if not empty
        unique_values = Read_from_attachment.get_datalist_for_update(claim_id)
        print('unique_values: ', unique_values)
        print('updating claim table with decoded values: ', unique_values)
        columns = ['project', 'part_name', 'part_id', 'customer']
        print('claim before update: ', Aggregation.select(claim_id))
        for i in range(0, 3):
            value = unique_values[i]
            column = columns[i]
            if value != None:
                Aggregation.update_db_claims(claim_id, value, column)
        print('claim after update: ', Aggregation.select(claim_id))


#test funkce dynamickeho pridavani obrazku
#path = 'C:/Users/START/PycharmProjects/databaze/claimfolder/claim2023_9_Material excess'

#Picture_resizing_for_qa.add_pictures_to_QA(path)

